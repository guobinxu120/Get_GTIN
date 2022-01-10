# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
from scrapy import signals
from xlrd import open_workbook

from openpyxl import Workbook, load_workbook
import numpy

class GetGtinPipeline(object):
	@classmethod
	def from_crawler(cls, crawler):
		pipeline = cls()
		# crawler.signals.connect(pipeline.spider_opened, signals.spider_opened)
		crawler.signals.connect(pipeline.spider_closed, signals.spider_closed)
		return pipeline
	def process_item(self, item, spider):
		return item
	def spider_closed(self, spider):
		data = spider.gtin_data_list
		real_value= []
		for d in spider.supplier_reference_list:
			try:
				idx = spider.key_list.index(d)
				real_value.append(spider.gtin_data_list[idx])
			except:
				real_value.append('')
		wb1 = UpdateWorkbook(r'Categories Managment.xlsx', worksheet=0)
		wb1['J2:J{0}'.format(str(len(real_value) + 1))] = real_value
		wb1.save()
		print data

class UpdateWorkbook(object):
	def __init__(self, fname, worksheet=0):
		self.fname = fname
		self.wb = load_workbook(fname)
		self.ws = self.wb.worksheets[worksheet]

	def save(self):
		self.wb.save(self.fname)

	def __setitem__(self, _range, values):
		"""
		 Assign Values to a Worksheet Range
		:param _range:  String e.g ['M6:M30']
		:param values: List: [row 1(col1, ... ,coln), ..., row n(col1, ... ,coln)]
		:return: None
		"""

		def _gen_value():
			for value in values:
				yield value

			if not isinstance(values, (list, numpy.ndarray)):
				raise ValueError('Values Type Error: Values have to be "list": values={}'.
								  format(type(values)))
			if isinstance(values, numpy.ndarray) and values.ndim > 1:
				raise ValueError('Values Type Error: Values of Type numpy.ndarray must have ndim=1; values.ndim={}'.
								  format(values.ndim))

		from openpyxl.utils import range_boundaries
		min_col, min_row, max_col, max_row = range_boundaries(_range)
		cols = ((max_col - min_col)+1)
		rows = ((max_row - min_row)+1)
		if cols * rows != len(values):
			raise ValueError('Number of List Values:{} does not match Range({}):{}'.
							 format(len(values), _range, cols * rows))

		value = _gen_value()
		n = 0
		for row_cells in self.ws.iter_rows(min_col=min_col, min_row=min_row,
										   max_col=max_col, max_row=max_row):
			for cell in row_cells:
				# cell.value = value.__next__()
				cell.value = values[n]
			n += 1


